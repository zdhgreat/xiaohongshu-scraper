"""SQLite + Markdown/CSV 渲染。

MVP 用 4 张表：notes / users / search_cache / crawl_state。
评论、图片、视频本地化延后到 P2。
"""

from __future__ import annotations

import csv
import json
import os
import re
import sqlite3
import sys
import threading
import time
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Iterable

from xhs_config import DB_PATH, MEDIA_DIR, OUTPUT_DIR, note_media_dir, sanitize_filename


def _atomic_write(path: Path, content: str, encoding: str = "utf-8") -> None:
    """原子写入文件：先写临时文件再 rename，防止崩溃时文件损坏。"""
    tmp = path.with_suffix(path.suffix + ".tmp")
    try:
        tmp.write_text(content, encoding=encoding)
        tmp.replace(path)
    except Exception:
        # replace 失败时清理临时文件
        try:
            tmp.unlink()
        except OSError:
            pass
        raise


SCHEMA = """
CREATE TABLE IF NOT EXISTS notes (
    note_id TEXT PRIMARY KEY,
    user_id TEXT,
    title TEXT,
    description TEXT,
    type TEXT,
    liked_count INTEGER DEFAULT 0,
    collected_count INTEGER DEFAULT 0,
    comment_count INTEGER DEFAULT 0,
    share_count INTEGER DEFAULT 0,
    ip_location TEXT,
    topics TEXT,
    published_at TEXT,
    xsec_token TEXT DEFAULT '',
    xsec_source TEXT DEFAULT '',
    video_url TEXT DEFAULT '',
    cover_url TEXT DEFAULT '',
    video_duration INTEGER DEFAULT 0,
    video_transcript TEXT DEFAULT '',
    video_ocr_text TEXT DEFAULT '',
    video_summary TEXT DEFAULT '',
    image_ocr_text TEXT DEFAULT '',
    image_summary TEXT DEFAULT '',
    image_mermaid TEXT DEFAULT '',
    raw_json TEXT,
    crawled_at TEXT DEFAULT CURRENT_TIMESTAMP
);

CREATE TABLE IF NOT EXISTS users (
    user_id TEXT PRIMARY KEY,
    nickname TEXT,
    avatar TEXT,
    description TEXT,
    fans_count INTEGER DEFAULT 0,
    follow_count INTEGER DEFAULT 0,
    notes_count INTEGER DEFAULT 0,
    location TEXT,
    raw_json TEXT,
    crawled_at TEXT DEFAULT CURRENT_TIMESTAMP
);

CREATE TABLE IF NOT EXISTS search_cache (
    keyword TEXT,
    page INTEGER,
    note_ids_json TEXT,
    crawled_at TEXT DEFAULT CURRENT_TIMESTAMP,
    PRIMARY KEY (keyword, page)
);

CREATE TABLE IF NOT EXISTS crawl_state (
    task_id TEXT PRIMARY KEY,
    task_type TEXT,
    target_id TEXT,
    cursor TEXT,
    status TEXT,
    last_error TEXT,
    updated_at TEXT DEFAULT CURRENT_TIMESTAMP
);

CREATE TABLE IF NOT EXISTS comments (
    comment_id TEXT PRIMARY KEY,
    note_id TEXT,
    parent_id TEXT DEFAULT '',
    user_id TEXT,
    nickname TEXT,
    content TEXT,
    like_count INTEGER DEFAULT 0,
    ip_location TEXT,
    pictures_json TEXT DEFAULT '[]',
    target_comment_id TEXT DEFAULT '',
    created_at TEXT,
    raw_json TEXT,
    crawled_at TEXT DEFAULT CURRENT_TIMESTAMP
);

CREATE INDEX IF NOT EXISTS idx_notes_user ON notes(user_id);
CREATE INDEX IF NOT EXISTS idx_notes_crawled ON notes(crawled_at);
CREATE INDEX IF NOT EXISTS idx_comments_note ON comments(note_id);
CREATE INDEX IF NOT EXISTS idx_comments_parent ON comments(parent_id);
"""


def _migrate(conn: sqlite3.Connection) -> None:
    """无损添加新字段，老 db 也兼容。"""
    cols = {r[1] for r in conn.execute("PRAGMA table_info(notes)").fetchall()}
    if "xsec_token" not in cols:
        conn.execute("ALTER TABLE notes ADD COLUMN xsec_token TEXT DEFAULT ''")
    if "xsec_source" not in cols:
        conn.execute("ALTER TABLE notes ADD COLUMN xsec_source TEXT DEFAULT ''")
    if "video_url" not in cols:
        conn.execute("ALTER TABLE notes ADD COLUMN video_url TEXT DEFAULT ''")
    if "cover_url" not in cols:
        conn.execute("ALTER TABLE notes ADD COLUMN cover_url TEXT DEFAULT ''")
    if "video_duration" not in cols:
        conn.execute("ALTER TABLE notes ADD COLUMN video_duration INTEGER DEFAULT 0")
    if "video_transcript" not in cols:
        conn.execute("ALTER TABLE notes ADD COLUMN video_transcript TEXT DEFAULT ''")
    if "video_ocr_text" not in cols:
        conn.execute("ALTER TABLE notes ADD COLUMN video_ocr_text TEXT DEFAULT ''")
    if "video_summary" not in cols:
        conn.execute("ALTER TABLE notes ADD COLUMN video_summary TEXT DEFAULT ''")
    if "image_ocr_text" not in cols:
        conn.execute("ALTER TABLE notes ADD COLUMN image_ocr_text TEXT DEFAULT ''")
    if "image_summary" not in cols:
        conn.execute("ALTER TABLE notes ADD COLUMN image_summary TEXT DEFAULT ''")
    if "image_mermaid" not in cols:
        conn.execute("ALTER TABLE notes ADD COLUMN image_mermaid TEXT DEFAULT ''")
    if "content_hash" not in cols:
        conn.execute("ALTER TABLE notes ADD COLUMN content_hash TEXT DEFAULT ''")
    # v2 对齐 PG schema 的增强列
    if "content" not in cols:
        conn.execute("ALTER TABLE notes ADD COLUMN content TEXT DEFAULT ''")
    if "note_url" not in cols:
        conn.execute("ALTER TABLE notes ADD COLUMN note_url TEXT DEFAULT ''")
    conn.commit()


_schema_done = False
_checkpoint_done = False
_init_lock = threading.Lock()


def db_retry(fn, *args, retries=3, delay=1.0, **kwargs):
    """执行 DB 操作，遇到 'database is locked' 自动重试。"""
    for i in range(retries):
        try:
            return fn(*args, **kwargs)
        except sqlite3.OperationalError as e:
            if "locked" in str(e).lower() and i < retries - 1:
                time.sleep(delay * (i + 1))
            else:
                raise


def connect() -> sqlite3.Connection:
    global _checkpoint_done, _schema_done
    DB_PATH.parent.mkdir(parents=True, exist_ok=True)
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    # busy_timeout: 多进程/多线程并发写冲突时排队等 30 秒
    conn.execute("PRAGMA busy_timeout = 30000")
    # WAL 模式允许并发读 + 串行写；先查再设，避免重复执行
    jm = conn.execute("PRAGMA journal_mode").fetchone()
    if jm is None or jm[0] != "wal":
        conn.execute("PRAGMA journal_mode=WAL")
    # 一次性初始化（checkpoint + DDL）加锁，防止多线程重复执行
    with _init_lock:
        if not _checkpoint_done:
            _checkpoint_done = True
            try:
                conn.execute("PRAGMA wal_checkpoint(TRUNCATE)")
            except Exception:
                pass
        if not _schema_done:
            _schema_done = True
            conn.executescript(SCHEMA)
            _migrate(conn)
    return conn


def _create_in_memory() -> sqlite3.Connection:
    """创建内存 SQLite 数据库（用于测试）。"""
    conn = sqlite3.connect(":memory:")
    conn.row_factory = sqlite3.Row
    conn.executescript(SCHEMA)
    _migrate(conn)
    return conn


def upsert_note(conn: sqlite3.Connection, note: dict[str, Any]) -> None:
    content_hash = note.get("content_hash", "")

    # 注意：新增字段时必须同步更新下方 ON CONFLICT DO UPDATE SET 子句
    cols = (
        "note_id user_id title description type liked_count collected_count "
        "comment_count share_count ip_location topics published_at "
        "xsec_token xsec_source video_url cover_url video_duration "
        "video_transcript video_ocr_text video_summary "
        "image_ocr_text image_summary image_mermaid raw_json content_hash "
        "content note_url"
    ).split()
    # content: 优先用 description，PG 同步时 hub_adapter 会从 raw_json.desc 丰富
    _content = note.get("description", "")
    _note_url = f"https://www.xiaohongshu.com/explore/{note.get('note_id', '')}"
    values = [
        note.get("note_id"),
        note.get("user_id"),
        note.get("title", ""),
        note.get("description", ""),
        note.get("type", "note"),
        int(note.get("liked_count", 0) or 0),
        int(note.get("collected_count", 0) or 0),
        int(note.get("comment_count", 0) or 0),
        int(note.get("share_count", 0) or 0),
        note.get("ip_location", ""),
        json.dumps(note.get("topics", []), ensure_ascii=False),
        note.get("published_at", ""),
        note.get("xsec_token", ""),
        note.get("xsec_source", ""),
        note.get("video_url", ""),
        note.get("cover_url", ""),
        int(note.get("video_duration", 0) or 0),
        note.get("video_transcript", ""),
        note.get("video_ocr_text", ""),
        note.get("video_summary", ""),
        note.get("image_ocr_text", ""),
        note.get("image_summary", ""),
        note.get("image_mermaid", ""),
        json.dumps(note.get("raw", {}), ensure_ascii=False),
        content_hash,
        _content,
        _note_url,
    ]
    # 防御性检查：确保 cols 数量与 values 数量一致
    assert len(cols) == len(values), f"upsert_note: cols({len(cols)}) != values({len(values)})"
    placeholders = ",".join("?" * len(cols))
    # 用 INSERT...ON CONFLICT，避免 INSERT OR REPLACE 把已有字段清成空
    # 遇到 database is locked 自动重试
    # 增量跳过：content_hash 相同时跳过更新（事务内判断，避免并发竞态）
    _skip_clause = ""
    if content_hash:
        _skip_clause = (
            f"CASE WHEN notes.content_hash = ? THEN notes.content_hash ELSE excluded.content_hash END, "
            f"crawled_at=CASE WHEN notes.content_hash = ? THEN notes.crawled_at ELSE CURRENT_TIMESTAMP END"
        )
        # 把 content_hash 值追加到 values 末尾供 CASE 参数使用
        values_with_skip = values + [content_hash, content_hash]
    else:
        _skip_clause = "excluded.content_hash, crawled_at=CURRENT_TIMESTAMP"
        values_with_skip = values

    # 去掉末尾的 content_hash=... 和 crawled_at=...，由 _skip_clause 控制
    _update_set = (
        f"user_id=excluded.user_id, title=excluded.title, description=excluded.description, "
        f"type=excluded.type, liked_count=excluded.liked_count, collected_count=excluded.collected_count, "
        f"comment_count=excluded.comment_count, share_count=excluded.share_count, "
        f"ip_location=excluded.ip_location, topics=excluded.topics, published_at=excluded.published_at, "
        f"xsec_token=CASE WHEN excluded.xsec_token != '' THEN excluded.xsec_token ELSE notes.xsec_token END, "
        f"xsec_source=CASE WHEN excluded.xsec_source != '' THEN excluded.xsec_source ELSE notes.xsec_source END, "
        f"video_url=CASE WHEN excluded.video_url != '' THEN excluded.video_url ELSE notes.video_url END, "
        f"cover_url=CASE WHEN excluded.cover_url != '' THEN excluded.cover_url ELSE notes.cover_url END, "
        f"video_duration=CASE WHEN excluded.video_duration > 0 THEN excluded.video_duration ELSE notes.video_duration END, "
        f"video_transcript=CASE WHEN excluded.video_transcript != '' THEN excluded.video_transcript ELSE notes.video_transcript END, "
        f"video_ocr_text=CASE WHEN excluded.video_ocr_text != '' THEN excluded.video_ocr_text ELSE notes.video_ocr_text END, "
        f"video_summary=CASE WHEN excluded.video_summary != '' THEN excluded.video_summary ELSE notes.video_summary END, "
        f"image_ocr_text=CASE WHEN excluded.image_ocr_text != '' THEN excluded.image_ocr_text ELSE notes.image_ocr_text END, "
        f"image_summary=CASE WHEN excluded.image_summary != '' THEN excluded.image_summary ELSE notes.image_summary END, "
        f"image_mermaid=CASE WHEN excluded.image_mermaid != '' THEN excluded.image_mermaid ELSE notes.image_mermaid END, "
        f"raw_json=CASE WHEN excluded.raw_json IS NOT NULL AND excluded.raw_json != '' AND (notes.raw_json IS NULL OR notes.raw_json = '' OR length(excluded.raw_json) > length(notes.raw_json)) THEN excluded.raw_json ELSE notes.raw_json END, "
        f"content=CASE WHEN excluded.content != '' THEN excluded.content ELSE notes.content END, "
        f"note_url=CASE WHEN excluded.note_url != '' THEN excluded.note_url ELSE COALESCE(notes.note_url, '') END, "
    )

    db_retry(
        conn.execute,
        f"INSERT INTO notes ({','.join(cols)}) VALUES ({placeholders}) "
        f"ON CONFLICT(note_id) DO UPDATE SET {_update_set} content_hash={_skip_clause}",
        values_with_skip,
    )
    conn.commit()


def update_video_analysis(
    conn: sqlite3.Connection,
    note_id: str,
    transcript: str = "",
    ocr_text: str = "",
    summary: str = "",
) -> None:
    """更新视频分析结果（转录 / OCR / 摘要），并刷新 crawled_at，不影响其他字段。"""
    conn.execute(
        "UPDATE notes SET video_transcript=?, video_ocr_text=?, video_summary=?, "
        "crawled_at=datetime('now') "
        "WHERE note_id=?",
        (transcript, ocr_text, summary, note_id),
    )
    conn.commit()


def update_image_analysis(
    conn: sqlite3.Connection,
    note_id: str,
    ocr_text: str = "",
    summary: str = "",
    mermaid: str = "",
) -> None:
    """更新图片分析结果（OCR / AI 描述 / Mermaid），不影响其他字段。"""
    conn.execute(
        "UPDATE notes SET image_ocr_text=?, image_summary=?, image_mermaid=? "
        "WHERE note_id=?",
        (ocr_text, summary, mermaid, note_id),
    )
    conn.commit()


def upsert_user(conn: sqlite3.Connection, user: dict[str, Any]) -> None:
    user_id = user.get("user_id")
    if not user_id:
        return
    conn.execute(
        """INSERT INTO users (user_id, nickname, avatar, description,
                              fans_count, follow_count, notes_count, location, raw_json)
           VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
           ON CONFLICT(user_id) DO UPDATE SET
               nickname = COALESCE(NULLIF(excluded.nickname, ''), users.nickname),
               avatar = COALESCE(NULLIF(excluded.avatar, ''), users.avatar),
               description = COALESCE(NULLIF(excluded.description, ''), users.description),
               fans_count = CASE WHEN excluded.fans_count > 0 THEN excluded.fans_count ELSE users.fans_count END,
               follow_count = CASE WHEN excluded.follow_count > 0 THEN excluded.follow_count ELSE users.follow_count END,
               notes_count = CASE WHEN excluded.notes_count > 0 THEN excluded.notes_count ELSE users.notes_count END,
               location = COALESCE(NULLIF(excluded.location, ''), users.location),
               raw_json = CASE WHEN excluded.raw_json IS NOT NULL AND excluded.raw_json != '' AND (users.raw_json IS NULL OR users.raw_json = '' OR length(excluded.raw_json) > length(users.raw_json)) THEN excluded.raw_json ELSE users.raw_json END
        """,
        (
            user_id,
            user.get("nickname", ""),
            user.get("avatar", ""),
            user.get("description", ""),
            int(user.get("fans_count", 0) or 0),
            int(user.get("follow_count", 0) or 0),
            int(user.get("notes_count", 0) or 0),
            user.get("location", ""),
            json.dumps(user.get("raw", {}), ensure_ascii=False),
        ),
    )
    conn.commit()


def save_search_page(
    conn: sqlite3.Connection, keyword: str, page: int, note_ids: list[str]
) -> None:
    conn.execute(
        "INSERT OR REPLACE INTO search_cache (keyword, page, note_ids_json) VALUES (?, ?, ?)",
        (keyword, page, json.dumps(note_ids, ensure_ascii=False)),
    )
    conn.commit()


def update_crawl_state(
    conn: sqlite3.Connection,
    task_id: str,
    task_type: str,
    target_id: str,
    cursor: str,
    status: str,
    last_error: str = "",
) -> None:
    conn.execute(
        """INSERT OR REPLACE INTO crawl_state
           (task_id, task_type, target_id, cursor, status, last_error, updated_at)
           VALUES (?, ?, ?, ?, ?, ?, ?)""",
        (task_id, task_type, target_id, cursor, status, last_error,
         datetime.now(timezone.utc).isoformat()),
    )
    conn.commit()


def get_crawl_state(conn: sqlite3.Connection, task_id: str) -> sqlite3.Row | None:
    cur = conn.execute("SELECT * FROM crawl_state WHERE task_id = ?", (task_id,))
    return cur.fetchone()


def get_note(conn: sqlite3.Connection, note_id: str) -> sqlite3.Row | None:
    cur = conn.execute("SELECT * FROM notes WHERE note_id = ?", (note_id,))
    return cur.fetchone()


def list_pending_corrections(conn: sqlite3.Connection, limit: int = 50) -> list[sqlite3.Row]:
    """待纠错视频笔记：有转录、有 OCR 参照、尚未标记纠错（video_summary 为空）。"""
    return conn.execute(
        "SELECT note_id, title, video_transcript, video_ocr_text FROM notes "
        "WHERE type='video' AND length(video_transcript) > 100 "
        "AND video_ocr_text != '' AND video_ocr_text != '[]' "
        "AND (video_summary = '' OR video_summary IS NULL) "
        "ORDER BY rowid DESC LIMIT ?",
        (limit,),
    ).fetchall()


def get_user(conn: sqlite3.Connection, user_id: str) -> sqlite3.Row | None:
    cur = conn.execute("SELECT * FROM users WHERE user_id = ?", (user_id,))
    return cur.fetchone()


def resolve_user_id(conn: sqlite3.Connection, name_or_id: str) -> str | None:
    """昵称 → user_id 查找。输入 user_id 则直接返回，输入昵称则查 DB。"""
    # 尝试作为 user_id 精确匹配
    row = conn.execute("SELECT user_id FROM users WHERE user_id = ?", (name_or_id,)).fetchone()
    if row:
        return row["user_id"]
    # 按昵称查找（精确优先，然后模糊）
    row = conn.execute("SELECT user_id FROM users WHERE nickname = ?", (name_or_id,)).fetchone()
    if row:
        return row["user_id"]
    row = conn.execute("SELECT user_id FROM users WHERE nickname LIKE ?", (f"%{name_or_id}%",)).fetchone()
    if row:
        return row["user_id"]
    return None


# ---------------------------------------------------------------------------
# Comments
# ---------------------------------------------------------------------------

def upsert_comment(conn: sqlite3.Connection, comment: dict[str, Any]) -> None:
    comment_id = comment.get("comment_id")
    if not comment_id:
        return
    conn.execute(
        """INSERT INTO comments (comment_id, note_id, parent_id, user_id, nickname,
                                 content, like_count, ip_location, pictures_json,
                                 target_comment_id, created_at, raw_json)
           VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
           ON CONFLICT(comment_id) DO UPDATE SET
               content = excluded.content,
               like_count = CASE WHEN excluded.like_count > 0 THEN excluded.like_count ELSE comments.like_count END,
               ip_location = COALESCE(NULLIF(excluded.ip_location, ''), comments.ip_location),
               pictures_json = excluded.pictures_json,
               raw_json = CASE WHEN excluded.raw_json IS NOT NULL AND excluded.raw_json != '' AND (comments.raw_json IS NULL OR comments.raw_json = '' OR length(excluded.raw_json) > length(comments.raw_json)) THEN excluded.raw_json ELSE comments.raw_json END
        """,
        (
            comment_id,
            comment.get("note_id"),
            comment.get("parent_id", ""),
            comment.get("user_id", ""),
            comment.get("nickname", ""),
            comment.get("content", ""),
            int(comment.get("like_count", 0) or 0),
            comment.get("ip_location", ""),
            json.dumps(comment.get("pictures", []), ensure_ascii=False),
            comment.get("target_comment_id", ""),
            comment.get("created_at", ""),
            json.dumps(comment.get("raw", {}), ensure_ascii=False),
        ),
    )
    conn.commit()


def iter_comments(conn: sqlite3.Connection, note_id: str) -> Iterable[sqlite3.Row]:
    """主评论按 like_count desc，子评论挂在主评论下面（按 created_at asc）"""
    main = conn.execute(
        "SELECT * FROM comments WHERE note_id = ? AND parent_id = '' "
        "ORDER BY like_count DESC, created_at ASC",
        (note_id,),
    ).fetchall()
    for m in main:
        yield m
        subs = conn.execute(
            "SELECT * FROM comments WHERE note_id = ? AND parent_id = ? "
            "ORDER BY created_at ASC",
            (note_id, m["comment_id"]),
        ).fetchall()
        yield from subs


def count_comments(conn: sqlite3.Connection, note_id: str) -> tuple[int, int]:
    """返回 (主评论数, 子评论数)"""
    main = conn.execute("SELECT COUNT(*) FROM comments WHERE note_id = ? AND parent_id = ''", (note_id,)).fetchone()[0]
    sub = conn.execute("SELECT COUNT(*) FROM comments WHERE note_id = ? AND parent_id != ''", (note_id,)).fetchone()[0]
    return main, sub


def iter_notes(
    conn: sqlite3.Connection, user_id: str | None = None
) -> Iterable[dict]:
    if user_id:
        cur = conn.execute(
            "SELECT * FROM notes WHERE user_id = ? ORDER BY published_at DESC",
            (user_id,),
        )
    else:
        cur = conn.execute("SELECT * FROM notes ORDER BY published_at DESC")
    for row in cur:
        yield dict(row)


def _note_context(conn: sqlite3.Connection, note_id: str, md_path: Path | None = None) -> dict:
    """提取渲染所需的公共数据，避免多个 _render_* 重复查库。"""
    note = get_note(conn, note_id)
    if not note:
        raise FileNotFoundError(f"note {note_id} not in DB")
    user = get_user(conn, note["user_id"]) if note["user_id"] else None
    raw = json.loads(note["raw_json"] or "{}")
    nickname = (user["nickname"] if user else "") or raw.get("user", {}).get("nickname", "")
    topics = json.loads(note["topics"] or "[]")
    topic_str = " ".join(f"#{t}" for t in topics) if topics else "—"

    # 计算从 MD 文件到 media 根目录的相对前缀
    media_rel_prefix = ".."
    if md_path is not None:
        try:
            depth = len(md_path.relative_to(OUTPUT_DIR).parts) - 1
            media_rel_prefix = "/".join([".."] * (depth + 1))
        except ValueError:
            pass

    media_dir = _find_media_dir(note_id, conn)
    local_images = sorted(media_dir.glob("img_*")) if media_dir.exists() else []
    local_video = next(iter(media_dir.glob("video.*")), None) if media_dir.exists() else None

    images = _extract_images(raw)
    if local_images:
        images = [str(Path(media_rel_prefix) / media_dir.relative_to(OUTPUT_DIR.parent) / p.name).replace("\\", "/") for p in local_images]
    video_url = _extract_video_url(raw)
    if local_video:
        video_url = str(Path(media_rel_prefix) / media_dir.relative_to(OUTPUT_DIR.parent) / local_video.name).replace("\\", "/")

    return {
        "note": note, "user": user, "raw": raw, "nickname": nickname,
        "topics": topics, "topic_str": topic_str, "media_dir": media_dir,
        "media_rel_prefix": media_rel_prefix,
        "images": images, "video_url": video_url,
        "local_video": local_video, "local_images": local_images,
    }


def _render_index(conn: sqlite3.Connection, ctx: dict) -> str:
    """渲染 index.md：元数据 + 正文 + 图片 + 视频基本信息。"""
    note = ctx["note"]
    lines = [
        f"# {note['title'] or '(无标题)'}",
        "",
        f"- **作者**: {ctx['nickname']} (@{note['user_id']})",
        f"- **发布时间**: {note['published_at'] or '—'}",
        f"- **IP属地**: {note['ip_location'] or '—'}",
        f"- **互动**: 赞 {note['liked_count']} | 藏 {note['collected_count']} | 评 {note['comment_count']} | 分享 {note['share_count']}",
        f"- **话题**: {ctx['topic_str']}",
        f"- **类型**: {note['type']}",
        f"- **笔记链接**: https://www.xiaohongshu.com/explore/{note['note_id']}",
        "",
        "## 正文",
        "",
        note["description"] or "(无正文)",
        "",
    ]
    if ctx["images"]:
        lines.append("## 图片")
        lines.append("")
        md_title = (note['title'] or '(无标题)').replace('[', '(').replace(']', ')')
        for i, url in enumerate(ctx["images"], 1):
            lines.append(f"![{md_title}·图{i}]({url})")
        lines.append("")

    # 视频基本信息（不含分析内容）
    if ctx["video_url"]:
        lines.append("## 视频")
        lines.append("")
        cover = note["cover_url"] or _extract_cover_url(ctx["raw"]) or ""
        if cover:
            local_cover = next(iter(ctx["media_dir"].glob("cover.*")), None) if ctx["media_dir"].exists() else None
            if local_cover:
                cover_ref = str(Path(ctx["media_rel_prefix"]) / ctx["media_dir"].relative_to(OUTPUT_DIR.parent) / local_cover.name).replace("\\", "/")
            else:
                cover_ref = cover
            lines.append(f"![封面]({cover_ref})")
            lines.append("")
        duration = note["video_duration"] or 0
        if duration:
            mins, secs = divmod(duration, 60)
            lines.append(f"- **时长**: {mins:02d}:{secs:02d}")
        lines.append(f"[视频链接]({ctx['video_url']})")
        if ctx["local_video"]:
            lines.append(f"- **本地文件**: `{ctx['local_video'].name}`")
        lines.append("")

    # 链接提示（指向其他子文件）
    links = []
    note_id = note["note_id"]
    transcript = note["video_transcript"] or ""
    ocr_text = note["video_ocr_text"] or ""
    summary = note["video_summary"] or ""
    image_summary = note["image_summary"] or ""
    image_ocr = note["image_ocr_text"] or ""
    image_mermaid = note["image_mermaid"] or ""
    main_n, _ = count_comments(conn, note_id)

    if summary or transcript or ocr_text:
        links.append("[视频分析](video.md)")
    if image_summary or image_ocr or image_mermaid:
        links.append("[图片分析](images.md)")
    if main_n:
        links.append(f"[评论](comments.md) ({main_n} 条)")

    if links:
        lines.append("---")
        lines.append("")
        for link in links:
            lines.append(f"- {link}")
        lines.append("")

    return "\n".join(lines)


def _render_video_section(ctx: dict) -> str | None:
    """渲染 video.md：摘要 + 转录 + OCR。无视频内容时返回 None。"""
    note = ctx["note"]
    transcript = note["video_transcript"] or ""
    ocr_text = note["video_ocr_text"] or ""
    summary = note["video_summary"] or ""

    if not (summary or transcript or ocr_text):
        return None

    title = note['title'] or '(无标题)'
    lines = [f"# {title} — 视频分析", ""]

    if summary:
        lines.append("## 摘要")
        lines.append("")
        lines.append(summary)
        lines.append("")
    elif transcript or ocr_text:
        lines.append("## 内容提取结果")
        lines.append("")
        lines.append(f"- 转录: {len(transcript)} 字")
        try:
            ocr_count = len(json.loads(ocr_text))
        except (json.JSONDecodeError, TypeError):
            ocr_count = 0
        lines.append(f"- 画面文字: {ocr_count} 帧")
        lines.append("")

    if transcript:
        lines.append("## 语音转录")
        lines.append("")
        lines.append(transcript)
        lines.append("")

    ocr_rendered = _render_video_ocr(ocr_text)
    if ocr_rendered:
        lines.append("## 画面文字")
        lines.append("")
        lines.append(ocr_rendered)
        lines.append("")

    return "\n".join(lines)


def _render_images_section(ctx: dict) -> str | None:
    """渲染 images.md：AI 描述 + OCR + Mermaid。无图片分析时返回 None。"""
    note = ctx["note"]
    image_summary = note["image_summary"] or ""
    image_ocr = note["image_ocr_text"] or ""
    image_mermaid = note["image_mermaid"] or ""

    if not (image_summary or image_ocr or image_mermaid):
        return None

    title = note['title'] or '(无标题)'
    lines = [f"# {title} — 图片分析", ""]

    if image_summary:
        lines.append("## AI 描述")
        lines.append("")
        lines.append(image_summary)
        lines.append("")
    if image_ocr:
        lines.append("## 图片文字")
        lines.append("")
        lines.append(image_ocr)
        lines.append("")
    if image_mermaid:
        lines.append("## 路线图 / 流程图")
        lines.append("")
        lines.append("```mermaid")
        lines.append(image_mermaid)
        lines.append("```")
        lines.append("")

    return "\n".join(lines)


def _render_comments_section(conn: sqlite3.Connection, note_id: str) -> str | None:
    """渲染 comments.md：全部评论。无评论时返回 None。"""
    main_n, sub_n = count_comments(conn, note_id)
    if not main_n:
        return None

    note = get_note(conn, note_id)
    title = (note["title"] or "(无标题)") if note else note_id
    lines = [f"# {title} — 评论", ""]
    lines.append(f"共 {main_n} 条主评论 + {sub_n} 条回复")
    lines.append("")

    for c in iter_comments(conn, note_id):
        indent = "" if c["parent_id"] == "" else "  "
        head = f"**{c['nickname']}** ({c['ip_location'] or '?'}, 赞 {c['like_count']})"
        lines.append(f"{indent}- {head}: {c['content']}")
    lines.append("")

    return "\n".join(lines)


def write_markdown_files(conn: sqlite3.Connection, note_id: str) -> list[Path]:
    """为单篇笔记生成子文件目录：index.md + video.md + images.md + comments.md。"""
    # 博主子目录
    note = get_note(conn, note_id)
    author_dir = OUTPUT_DIR
    if note and note["user_id"]:
        user = get_user(conn, note["user_id"])
        nickname = (user["nickname"] if user else "") or ""
        if nickname:
            author_dir = OUTPUT_DIR / sanitize_filename(nickname, 30)

    # 笔记子目录
    slug = _human_filename(conn, note_id, "")
    note_dir = author_dir / slug
    note_dir.mkdir(parents=True, exist_ok=True)

    # 公共数据
    index_path = note_dir / "index.md"
    ctx = _note_context(conn, note_id, md_path=index_path)

    files = []

    # index.md（必有）
    index_content = _render_index(conn, ctx)
    _atomic_write(index_path, index_content)
    files.append(index_path)

    # video.md（有视频分析时生成）
    video_content = _render_video_section(ctx)
    if video_content:
        path = note_dir / "video.md"
        _atomic_write(path, video_content)
        files.append(path)

    # images.md（有图片分析时生成）
    images_content = _render_images_section(ctx)
    if images_content:
        path = note_dir / "images.md"
        _atomic_write(path, images_content)
        files.append(path)

    # comments.md（有评论时生成）
    comments_content = _render_comments_section(conn, note_id)
    if comments_content:
        path = note_dir / "comments.md"
        _atomic_write(path, comments_content)
        files.append(path)

    return files


def _render_video_ocr(ocr_text: str) -> str:
    """智能渲染视频 OCR：JSON 解析 → 去重 → 噪声过滤。旧数据纯文本直接返回。"""
    if not ocr_text:
        return ""
    try:
        items = json.loads(ocr_text)
    except (json.JSONDecodeError, TypeError):
        return ocr_text  # 旧数据兼容

    seen = set()
    lines = []
    for item in items:
        text = item.get("text", "").strip()
        if not text or len(text) < 2 or text in seen:
            continue
        seen.add(text)
        time_label = item.get("time", "")
        if time_label:
            lines.append(f"- [{time_label}] {text}")
        else:
            lines.append(f"- {text}")
    return "\n".join(lines) if lines else ""


def write_markdown(conn: sqlite3.Connection, note_id: str) -> Path:
    """导出笔记为 MD 子文件目录，返回 index.md 路径。"""
    files = write_markdown_files(conn, note_id)
    return files[0]  # index.md


def _human_filename(conn: sqlite3.Connection, note_id: str, ext: str) -> str:
    """生成人类可读的文件名：{标题}_{note_id前8位}.ext

    无标题时退化到 {note_id}.ext。
    """
    row = get_note(conn, note_id)
    if not row:
        return f"{note_id}{ext}"
    title = row["title"] or ""
    if not title:
        desc = row["description"] or ""
        title = desc[:50] if desc else ""
    # 清理标题：去掉文件名不合法字符
    title = re.sub(r'[<>:"/\\|?*\n\r\t]', '', title)
    title = title.replace(' ', '_')
    title = title[:70]  # 截断
    if not title:
        return f"{note_id}{ext}"
    return f"{title}_{note_id[:8]}{ext}"


def render_update_summary(conn: sqlite3.Connection, note_id: str) -> str:
    """生成笔记当前状态摘要，用于更新通知。"""
    row = get_note(conn, note_id)
    if not row:
        return ""
    parts = []
    title = row["title"] or "(无标题)"

    # 基本信息
    type_label = "视频" if row["type"] == "video" else "图文"
    parts.append(f"《{title}》({type_label})")

    # 媒体状态
    media_dir = _find_media_dir(note_id, conn)
    local_imgs = len(list(media_dir.glob("img_*"))) if media_dir.exists() else 0
    local_video = next(iter(media_dir.glob("video.*")), None) is not None if media_dir.exists() else False
    media_items = []
    if local_imgs:
        media_items.append(f"{local_imgs} 张图片")
    if local_video:
        media_items.append("1 个视频")
    if media_items:
        parts.append("已下载: " + "、".join(media_items))

    # 分析状态
    analysis_items = []
    if row["video_transcript"]:
        analysis_items.append(f"语音转录 {len(row['video_transcript'])} 字")
    if row["video_ocr_text"]:
        analysis_items.append(f"画面文字 {len(row['video_ocr_text'])} 字")
    if row["video_summary"]:
        analysis_items.append("AI 摘要")
    if analysis_items:
        parts.append("视频分析: " + "、".join(analysis_items))

    # 图片分析状态
    image_items = []
    if row["image_ocr_text"]:
        image_items.append(f"图片OCR {len(row['image_ocr_text'])} 字")
    if row["image_summary"]:
        image_items.append("图片AI描述")
    if row["image_mermaid"]:
        image_items.append("路线图/流程图")
    if image_items:
        parts.append("图片分析: " + "、".join(image_items))

    # 评论
    main_n, sub_n = count_comments(conn, note_id)
    if main_n:
        parts.append(f"{main_n} 条评论")

    return " | ".join(parts)


CSV_HEADERS = [
    "序号", "标题", "作者", "发布时间", "类型",
    "点赞", "收藏", "评论", "分享", "IP属地",
    "话题", "笔记链接", "媒体", "正文摘要",
]


def write_csv(conn: sqlite3.Connection, path: Path | None = None, user_id: str | None = None) -> list[Path]:
    """导出 CSV。按博主分文件，放到 output/<博主名>/ 目录。返回生成的文件列表。"""
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

    # 按 user_id 分组
    if user_id:
        groups = {user_id: list(iter_notes_by_user(conn, user_id))}
    else:
        groups: dict[str, list] = {}
        for note in iter_notes(conn):
            uid = note["user_id"] or "_unknown"
            groups.setdefault(uid, []).append(note)

    files = []
    for uid, notes in groups.items():
        # 确定博主目录
        user = get_user(conn, uid) if uid != "_unknown" else None
        nickname = (user["nickname"] if user else "") or uid
        safe_author = sanitize_filename(nickname, 30)
        author_dir = OUTPUT_DIR / safe_author
        author_dir.mkdir(parents=True, exist_ok=True)

        csv_path = path or (author_dir / f"{safe_author}_笔记列表.csv")
        with open(csv_path, "w", encoding="utf-8-sig", newline="") as f:
            writer = csv.writer(f, quoting=csv.QUOTE_ALL)
            writer.writerow(CSV_HEADERS)
            for i, note in enumerate(notes, 1):
                note_user = get_user(conn, note["user_id"]) if note["user_id"] else None
                topics = json.loads(note["topics"] or "[]")
                desc = note["description"] or ""

                # 媒体摘要
                csv_media_dir = _find_media_dir(note["note_id"], conn)
                media_parts = []
                if csv_media_dir.exists():
                    img_count = len(list(csv_media_dir.glob("img_*")))
                    has_video = any(csv_media_dir.glob("video.*"))
                    if has_video:
                        media_parts.append("视频")
                    if img_count:
                        media_parts.append(f"{img_count}张图")
                media_str = "+".join(media_parts) if media_parts else "—"

                # 正文摘要：截断 + 换行替换
                summary = desc[:200].replace("\n", " ").replace("\r", " ").strip()

                writer.writerow([
                    i,
                    note["title"] or "",
                    (note_user["nickname"] if note_user else "") or "",
                    note["published_at"] or "",
                    "视频" if note["type"] == "video" else "图文",
                    note["liked_count"],
                    note["collected_count"],
                    note["comment_count"],
                    note["share_count"],
                    note["ip_location"] or "",
                    " ".join(f"#{t}" for t in topics),
                    f"https://www.xiaohongshu.com/explore/{note['note_id']}",
                    media_str,
                    summary,
                ])
        files.append(csv_path)
        path = None  # 后续分组自动生成路径

    return files


def iter_notes_by_user(conn: sqlite3.Connection, user_id: str):
    """按 user_id 迭代笔记。"""
    cur = conn.execute(
        "SELECT * FROM notes WHERE user_id = ? ORDER BY published_at DESC",
        (user_id,),
    )
    for row in cur:
        yield dict(row)


def write_json(conn: sqlite3.Connection, path: Path | None = None, user_id: str | None = None) -> Path:
    """导出笔记为 JSON 文件。user_id 过滤指定博主。"""
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    date_str = datetime.now().strftime('%Y%m%d_%H%M%S')
    if user_id:
        user = get_user(conn, user_id)
        nickname = (user["nickname"] if user else "") or user_id
        safe_author = sanitize_filename(nickname, 30)
        author_dir = OUTPUT_DIR / safe_author
        author_dir.mkdir(parents=True, exist_ok=True)
        path = path or (author_dir / f"{safe_author}_笔记.json")
    else:
        path = path or (OUTPUT_DIR / f"xhs_export_{date_str}.json")
    notes = []
    source = iter_notes_by_user(conn, user_id) if user_id else iter_notes(conn)
    for note in source:
        d = dict(note)
        d["topics"] = json.loads(d.get("topics") or "[]")
        d["raw_json"] = json.loads(d.get("raw_json") or "{}")
        notes.append(d)
    _atomic_write(path, json.dumps(notes, ensure_ascii=False, indent=2, default=str))
    return path


def write_xlsx(conn: sqlite3.Connection, path: Path | None = None, user_id: str | None = None) -> Path:
    """导出笔记为 XLSX 文件（多 sheet）。user_id 过滤指定博主。需要 openpyxl。"""
    try:
        import openpyxl  # type: ignore
    except ImportError:
        print("[ERR] 导出 xlsx 需要 openpyxl。pip install openpyxl", file=sys.stderr)
        raise
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    if user_id:
        user = get_user(conn, user_id)
        nickname = (user["nickname"] if user else "") or user_id
        safe_author = sanitize_filename(nickname, 30)
        author_dir = OUTPUT_DIR / safe_author
        author_dir.mkdir(parents=True, exist_ok=True)
        date_str = datetime.now().strftime('%Y%m%d_%H%M%S')
        path = path or (author_dir / f"{safe_author}_笔记.xlsx")
    else:
        date_str = datetime.now().strftime('%Y%m%d_%H%M%S')
        path = path or (OUTPUT_DIR / f"xhs_export_{date_str}.xlsx")
    wb = openpyxl.Workbook()

    # Sheet 1: Notes
    ws_notes = wb.active
    ws_notes.title = "Notes"
    note_cols = [
        "note_id", "user_id", "title", "description", "type",
        "liked_count", "collected_count", "comment_count", "share_count",
        "ip_location", "published_at", "xsec_token",
        "video_url", "cover_url", "video_duration",
        "video_summary", "image_ocr_text", "image_summary",
        "crawled_at",
    ]
    ws_notes.append(note_cols)
    for note in (iter_notes_by_user(conn, user_id) if user_id else iter_notes(conn)):
        ws_notes.append([note.get(c, "") for c in note_cols])

    # Sheet 2: Users
    ws_users = wb.create_sheet("Users")
    user_cols = ["user_id", "nickname", "avatar", "fans_count", "notes_count", "desc"]
    ws_users.append(user_cols)
    for row in conn.execute("SELECT * FROM users ORDER BY fans_count DESC"):
        d = dict(row)
        ws_users.append([d.get(c, "") for c in user_cols])

    # Sheet 3: Comments
    ws_comments = wb.create_sheet("Comments")
    comment_cols = ["comment_id", "note_id", "content", "nickname", "like_count", "created_at"]
    ws_comments.append(comment_cols)
    for row in conn.execute("SELECT * FROM comments ORDER BY created_at DESC"):
        d = dict(row)
        ws_comments.append([d.get(c, "") for c in comment_cols])

    wb.save(path)
    return path


def _extract_images(raw: dict) -> list[str]:
    images: list[str] = []
    for img in raw.get("image_list", []) or []:
        url = img.get("url_default") or img.get("url") or ""
        info_list = img.get("info_list") or []
        for info in info_list:
            if info.get("image_scene") in ("WB_DFT", "WB_PRV"):
                url = info.get("url") or url
                break
        if url:
            images.append(url)
    return images


def _extract_video_url(raw: dict) -> str | None:
    video = raw.get("video") or {}
    media = video.get("media") or {}
    stream = media.get("stream") or {}
    for codec_key in ("h264", "h265", "av1"):
        items = stream.get(codec_key) or []
        if items:
            return items[0].get("master_url") or items[0].get("backup_urls", [None])[0]
    return None


def _extract_cover_url(raw: dict) -> str:
    """从 raw_json 提取视频封面图 URL。"""
    video = raw.get("video") or {}
    # 尝试多种路径
    for path in [
        lambda: video.get("cover", ""),
        lambda: (video.get("image_list") or [{}])[0].get("url_default", "") if video.get("image_list") else "",
        lambda: (raw.get("image_list") or [{}])[0].get("url_default", "") if raw.get("image_list") else "",
    ]:
        try:
            url = path()
            if url:
                return url
        except (IndexError, TypeError):
            continue
    return ""


def _extract_video_duration(raw: dict) -> int:
    """从 raw_json 提取视频时长（秒）。"""
    video = raw.get("video") or {}
    dur = video.get("duration")
    if dur:
        return int(dur) // 1000 if int(dur) > 1000 else int(dur)
    return 0


def _find_media_dir(note_id: str, conn: sqlite3.Connection) -> Path:
    """查找笔记的本地媒体目录（兼容新旧两种结构）。

    新结构: media/<博主名>/<笔记标题>/
    旧结构: media/<note_id>/

    当新旧目录都存在时，优先返回有实际内容的那个；
    如果都有内容，优先新路径（后续下载会写入新路径）。
    """
    new_dir = note_media_dir(note_id, conn)
    new_has_content = new_dir.exists() and any(new_dir.iterdir())
    legacy_dir = MEDIA_DIR / note_id
    legacy_has_content = legacy_dir.exists() and any(legacy_dir.iterdir())

    if new_has_content:
        return new_dir
    if legacy_has_content:
        return legacy_dir
    # 都不存在或都为空，返回新路径（作为目标目录）
    return new_dir
